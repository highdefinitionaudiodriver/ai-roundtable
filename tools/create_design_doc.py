from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "AI_Roundtable_設計書.xlsx"


HEADER = PatternFill("solid", fgColor="087F8C")
SUBHEADER = PatternFill("solid", fgColor="E8F4F3")
THIN = Side(style="thin", color="D9D3CA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = Font(color="FFFFFF", bold=True)


def add_table(ws, headers: list[str], rows: list[list[str]], widths: list[int]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws, widths)


wb = Workbook()
ws = wb.active
ws.title = "概要"
add_table(
    ws,
    ["項目", "内容"],
    [
        ["システム名", "AI Roundtable"],
        ["目的", "あるAIの回答を、他のAIに共有し、レビュー・反論・補足・統合を行う。"],
        ["対象", "PCブラウザ、スマホPWA、対応ブラウザの共有メニュー。"],
        ["主要機能", "AI回答貼り付け、共有プロンプト生成、AI別リンク、PWA、共有ターゲット、自動議論モード。"],
        ["対応言語", "入力言語を基本として同言語で応答。UIは日本語、英語、中国語、韓国語、スペイン語、フランス語の切替を想定。"],
        ["API連携", "OpenAI、Anthropic Claude、Google Gemini。APIキー未設定時はプロンプトコピー運用。"],
        ["設計方針", "サーバー側で無限ループを持たず、クライアントが1発言ずつAPIを呼び停止制御する。自動議論は10ターンまたは10分で停止する。"],
        ["透明性", "アプリ内実行は選択した同一AI/同一モデルが複数視点を演じる方式であり、複数社AIによる独立検証ではない。"],
    ],
    [24, 90],
)

ws = wb.create_sheet("機能一覧")
add_table(
    ws,
    ["ID", "機能", "説明", "実装ファイル"],
    [
        ["F-01", "元AI回答入力", "質問と回答を貼り付ける。共有URLからの自動入力にも対応。", "public/index.html, public/app.js"],
        ["F-02", "共有プロンプト生成", "他AIへ投入しやすいレビュー依頼プロンプトを生成しクリップボードへコピー。", "public/app.js"],
        ["F-03", "AIリンク", "ChatGPT、Claude、Gemini、Perplexity、Grok、Copilotを別タブで開く。", "public/app.js"],
        ["F-04", "通常検討会", "レビュアー、反論役、補足役、統合役の4発言を生成。", "server.py"],
        ["F-05", "自動議論モード", "開始後、停止ボタン押下まで役割を巡回して発言生成を継続。", "public/app.js, server.py"],
        ["F-06", "停止制御", "停止ボタンで次回以降の発言生成を止める。実行中リクエストは完了後に停止。", "public/app.js"],
        ["F-07", "PWA", "manifest、service worker、アイコンによりホーム画面追加に対応。", "public/manifest.webmanifest, public/sw.js"],
        ["F-08", "共有ターゲット", "対応環境で共有テキスト・URLを取り込みフォームへ反映。", "public/manifest.webmanifest, public/app.js"],
        ["F-09", "安全ガード", "明確な危険カテゴリはブロックし、高リスクカテゴリは確認モーダルを表示。", "public/app.js, public/index.html"],
        ["F-10", "CSRF/Origin保護", "ローカルAPIへのPOSTは同一OriginとCSRFトークンを必須化。", "server.py, public/app.js"],
        ["F-11", "透明性表示", "同一モデルのロールプレイであり独立検証ではない旨を常時表示。", "public/index.html, public/app.js"],
        ["F-12", "外部取り込み警告", "URL/共有メニューから取り込まれた内容は警告バナーを表示。", "public/app.js, public/index.html"],
        ["F-13", "プロバイダ上限", "各プロバイダ呼び出しに出力トークン上限を明示。Geminiは安全設定も指定。", "server.py"],
    ],
    [12, 24, 70, 44],
)

ws = wb.create_sheet("画面設計")
add_table(
    ws,
    ["画面/領域", "項目", "UI", "動作"],
    [
        ["トップ", "言語選択", "select", "Autoまたは指定言語でUI文言を切替。"],
        ["入力パネル", "元の質問・相談", "textarea", "ユーザーの元質問または共有元URLを入力。"],
        ["入力パネル", "このAIに言われた内容", "textarea", "他AIからの回答本文。必須。"],
        ["入力パネル", "議論モード", "select", "balanced / strict / creative / expert を選択。"],
        ["入力パネル", "実行AI", "select", "openai / anthropic / gemini を選択。"],
        ["入力パネル", "検討会を実行", "button", "4役の通常検討会を実行。"],
        ["入力パネル", "自動議論を開始", "button", "停止まで1発言ずつ継続生成。開始中は入力をロック。"],
        ["入力パネル", "停止", "button", "自動議論の継続フラグをOFFにする。"],
        ["確認モーダル", "高リスク確認", "dialog", "医療、法律、金融等を検知した場合、カテゴリ別ガイダンスと送信先プロバイダを表示。"],
        ["警告バナー", "外部取り込み", "section", "URLまたは共有メニューで入力された内容であることを明示。"],
        ["共有パネル", "AIリンク", "link buttons", "プロンプトをコピーしたうえで各AIサービスを開く。"],
        ["結果パネル", "タイムライン", "articles", "役割名と本文を時系列表示。"],
    ],
    [22, 26, 22, 70],
)

ws = wb.create_sheet("API設計")
add_table(
    ws,
    ["エンドポイント", "メソッド", "入力", "出力", "用途"],
    [
        ["/api/providers", "GET", "なし", "providers: {openai, anthropic, gemini}", "APIキー設定状況の確認。"],
        ["/api/roundtable", "POST", "question, sourceAnswer, mode, provider", "transcript[]", "通常検討会を一括実行。"],
        ["/api/roundtable-step", "POST", "question, sourceAnswer, mode, provider, role, transcript[]", "entry", "自動議論の1発言生成。"],
    ],
    [28, 14, 48, 44, 54],
)

ws = wb.create_sheet("データ設計")
add_table(
    ws,
    ["名称", "型", "説明"],
    [
        ["question", "string", "元のユーザー質問。任意。"],
        ["sourceAnswer", "string", "他AIの回答本文。必須。"],
        ["mode", "string", "議論モード。balanced, strict, creative, expert。"],
        ["provider", "string", "実行AI。openai, anthropic, gemini。"],
        ["transcript", "array", "発言履歴。role/textの配列。"],
        ["role", "string", "reviewer, skeptic, expander, moderator, synthesizer。"],
        ["entry", "object", "1発言。role/text。"],
    ],
    [26, 18, 80],
)

ws = wb.create_sheet("自動議論フロー")
add_table(
    ws,
    ["順序", "処理", "詳細"],
    [
        ["1", "開始", "ユーザーが自動議論を開始。sourceAnswer必須チェック。"],
        ["2", "状態更新", "autoRunning=true、transcript初期化、入力UIをロック。"],
        ["3", "役割選択", "reviewer -> skeptic -> expander -> moderator -> synthesizer を巡回。"],
        ["4", "1発言生成", "/api/roundtable-step に現在の履歴と次ロールを送信。"],
        ["5", "表示更新", "戻ったentryをtimelineへ追加。"],
        ["6", "継続判定", "autoRunningがtrue、10ターン未満、10分未満なら次発言へ。停止押下でfalse。"],
        ["7", "停止", "入力UIを解除し、停止または上限到達ステータスを表示。"],
    ],
    [10, 26, 90],
)

ws = wb.create_sheet("安全ガード")
add_table(
    ws,
    ["区分", "対象例", "アプリ挙動", "補足"],
    [
        ["禁止", "自傷助長、違法行為、武器、詐欺、他者危害、秘密情報/個人情報パターン", "実行前にブロック", "クライアント側とサーバ側の両方で拒否。"],
        ["禁止時支援", "自傷カテゴリ", "実行はブロックしつつ相談先ガイダンスを表示", "冷たい拒否だけにせず、身近な人・医療機関・緊急窓口への相談を促す。"],
        ["高リスク", "メンタルヘルス支援、医療、法律、金融、採用、教育、住宅、行政", "確認モーダルを表示", "カテゴリ別ガイダンス、相談先の考え方、送信先プロバイダを明示。"],
        ["低リスク", "旅行、学習、一般的な文章検討、コードレビュー", "通常実行", "必要に応じてAPIキー未設定時はプロンプトコピーへ誘導。"],
        ["自動議論", "高リスクカテゴリ検知時", "追加警告を表示", "断定的な内容の増幅とAPI費用に注意。"],
    ],
    [18, 44, 32, 70],
)

ws = wb.create_sheet("非機能・注意事項")
add_table(
    ws,
    ["分類", "内容"],
    [
        ["セキュリティ", "APIキーは環境変数で管理し、フロントエンドへ露出しない。"],
        ["セキュリティ", "POST APIは同一OriginとCSRFトークンを検証し、任意サイトからのローカルAPI濫用を抑止する。"],
        ["セキュリティ", "127.0.0.1バインドの個人ローカル利用を前提とする。同一端末上の別プロセスは到達可能なため、信頼できる端末で実行する。"],
        ["プライバシー", "元AI回答や質問には機密情報が含まれうるため、外部AI送信前にユーザー確認導線を強化する余地あり。"],
        ["プライバシー", "秘密情報・個人情報パターンは既定ブロック。URL/共有メニュー取り込み時は警告バナーを表示。"],
        ["可用性", "APIキー未設定または通信失敗時は、プロンプトコピー運用へフォールバック。"],
        ["制限", "自動議論は10ターンまたは10分で自動停止。API費用と利用制限に注意。"],
        ["制限", "OpenAI/Anthropic/Geminiに出力上限を設定。Geminiには安全設定を明示。"],
        ["透明性", "アプリ内の複数役割は選択した単一プロバイダ/モデルのロールプレイであり、独立した複数AIの合議ではない。"],
        ["高リスク領域", "医療、法律、金融、安全などは専門家確認を促すシステムプロンプトを付与。"],
        ["スマホ", "PWAの共有ターゲット対応はOS/ブラウザに依存。iPhoneはホーム画面追加とコピー運用を推奨。"],
    ],
    [22, 100],
)

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = SUBHEADER

wb.save(OUT)
print(OUT)
